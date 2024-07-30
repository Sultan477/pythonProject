import os
import openpyxl
from flet import (
    Page, TextField, ElevatedButton, Column, Row, Container, Text, AlertDialog, app
)

class Student:
    def __init__(self, first_name, last_name, grade, address, school_info, student_phone, school_phone):
        self.first_name = first_name
        self.last_name = last_name
        self.grade = grade
        self.address = address
        self.school_info = school_info
        self.student_phone = student_phone
        self.school_phone = school_phone

class SchoolDatabase:
    def __init__(self, file_name="student_data.xlsx"):
        self.file_name = file_name
        if not os.path.exists(file_name):
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self._initialize_sheet()
        else:
            self.workbook = openpyxl.load_workbook(file_name)
            self.sheet = self.workbook.active

    def _initialize_sheet(self):
        headers = ["First Name", "Last Name", "Grade", "Address", "School Info", "Student Phone", "School Phone"]
        self.sheet.append(headers)
        self.workbook.save(self.file_name)

    def add_student(self, student):
        self.sheet.append([student.first_name, student.last_name, student.grade, student.address, student.school_info, student.student_phone, student.school_phone])
        self.workbook.save(self.file_name)

    def search_students_by_grade(self, grade):
        return [row for row in self.sheet.iter_rows(values_only=True) if row[2] == grade]

    def search_students_by_name(self, name):
        return [row for row in self.sheet.iter_rows(values_only=True) if row[0] == name or row[1] == name]

def main(page: Page):
    database = SchoolDatabase()

    input_first_name = TextField(label="First Name")
    input_last_name = TextField(label="Last Name")
    input_grade = TextField(label="Grade")
    input_street = TextField(label="Street")
    input_city = TextField(label="City")
    input_zip = TextField(label="ZIP Code")
    input_school_number = TextField(label="School Number")
    input_teacher_last_name = TextField(label="Teacher Last Name")
    input_student_phone_type = TextField(label="Student Phone Type")
    input_student_phone_number = TextField(label="Student Phone Number")
    input_school_phone_type = TextField(label="School Phone Type")
    input_school_phone_number = TextField(label="School Phone Number")

    add_student_dialog = AlertDialog(
        title=Text("Add Student"),
        content=Column([
            Row([
                Column([
                    input_first_name,
                    input_last_name,
                    input_grade,
                    input_street,
                    input_city,
                    input_zip
                ], expand=True),
                Column([
                    input_school_number,
                    input_teacher_last_name,
                    input_student_phone_type,
                    input_student_phone_number,
                    input_school_phone_type,
                    input_school_phone_number
                ], expand=True)
            ])
        ]),
        actions=[
            ElevatedButton(text="Save", on_click=lambda e: add_student(e)),
            ElevatedButton(text="Cancel", on_click=lambda e: close_add_student_dialog(e))
        ]
    )

    def show_add_student_dialog(e):
        page.dialog = add_student_dialog
        add_student_dialog.open = True
        page.update()

    def close_add_student_dialog(e):
        add_student_dialog.open = False
        page.update()

    def add_student(e):
        student = Student(
            first_name=input_first_name.value,
            last_name=input_last_name.value,
            grade=input_grade.value,
            address=f"{input_street.value}, {input_city.value}, {input_zip.value}",
            school_info=f"{input_school_number.value}, {input_teacher_last_name.value}",
            student_phone=f"{input_student_phone_type.value}: {input_student_phone_number.value}",
            school_phone=f"{input_school_phone_type.value}: {input_school_phone_number.value}"
        )
        database.add_student(student)
        close_add_student_dialog(e)

    search_input_grade = TextField(label="Grade")
    search_input_name = TextField(label="Name")
    search_results_column = Column()

    def search_by_grade(e):
        results = database.search_students_by_grade(search_input_grade.value)
        search_results_column.controls.clear()
        for result in results:
            search_results_column.controls.append(Text(", ".join(result)))
        page.update()

    def search_by_name(e):
        results = database.search_students_by_name(search_input_name.value)
        search_results_column.controls.clear()
        for result in results:
            search_results_column.controls.append(Text(", ".join(result)))
        page.update()

    page.add(
        Column([
            ElevatedButton(text="Add Student", on_click=show_add_student_dialog),
            Container(height=10),
            Row([
                search_input_grade,
                ElevatedButton(text="Search by Grade", on_click=search_by_grade)
            ]),
            Row([
                search_input_name,
                ElevatedButton(text="Search by Name", on_click=search_by_name)
            ]),
            Container(height=10),
            Text("Search Results:"),
            search_results_column
        ])
    )

app(target=main)