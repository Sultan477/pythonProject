from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Создание нового файла Excel
wb = Workbook()

# Получение активного листа
ws = wb.active

# Запись данных в ячейку
ws['A1'] = 'Hello, World!'

# Применение стиля к ячейке
ws['A1'].font = Font(name='Arial', size=14, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Запись данных в диапазон ячеек
data = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]

for row in data:
    ws.append(row)

# Сохранение файла с изменениями
wb.save('styled_example.xlsx')