# Импортируем необходимые модули из библиотеки openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# Создание нового файла Excel
wb = Workbook()

# Получение активного листа
ws = wb.active

# Запись данных в ячейку
ws['A1'] = 'Hello, World!'

# Применение стиля к ячейке
ws['A1'].font = Font(name='Arial', size=14, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Запись данных в диапазон ячеек
data = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]

for row in data:
    ws.append(row)

# Сохранение файла с изменениями
wb.save('styled_example.xlsx')

# Открытие существующего файла Excel
wb2 = load_workbook('styled_example.xlsx')

# Получение активного листа
ws2 = wb2.active

# Чтение данных из ячейки A1
value = ws2['A1'].value
print("Value in A1:", value)

# Чтение диапазона ячеек
print("Reading range A1:C3:")
for row in ws2['A1:C3']:
    for cell in row:
        print(cell.value)

# Создание нового листа
ws3 = wb2.create_sheet("NewSheet")

# Переименование листа
ws3.title = "RenamedSheet"

# Запись данных в новую ячейку на новом листе
ws3['A1'] = "New Sheet Data"

# Сохранение файла с изменениями
wb2.save('updated_example.xlsx')